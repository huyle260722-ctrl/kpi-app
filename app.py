from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import datetime
import io, os

app = Flask(__name__, static_folder='static', static_url_path='')
CORS(app)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'TEMPLATE_KPI.xlsx')

@app.route('/')
def index():
    return app.send_static_file('index.html')

@app.route('/api/export', methods=['POST'])
def export_excel():
    payload = request.get_json()
    year_month  = payload.get('month', '')
    report_data = payload.get('data', {})
    if not year_month or not report_data:
        return jsonify({'error': 'Thiếu dữ liệu'}), 400
    try:
        wb = build_from_template(report_data, year_month)
        buf = io.BytesIO()
        wb.save(buf); buf.seek(0)
        y, m = year_month.split('-')
        return send_file(buf, as_attachment=True,
                         download_name=f'Bao_Cao_Thang_{m}_{y}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Helpers ─────────────────────────────────────────────────────
def get_week(day):
    if day <= 7:  return 1
    if day <= 14: return 2
    if day <= 21: return 3
    return 4

def copy_style(src, dst):
    """Copy toàn bộ style từ cell src sang cell dst."""
    if src.font:        dst.font        = copy(src.font)
    if src.fill:        dst.fill        = copy(src.fill)
    if src.border:      dst.border      = copy(src.border)
    if src.alignment:   dst.alignment   = copy(src.alignment)
    if src.number_format: dst.number_format = src.number_format

def copy_row_style(ws_tpl, tpl_row, ws_out, out_row, num_cols=12):
    """Clone style của tpl_row sang out_row."""
    for col in range(1, num_cols + 1):
        src = ws_tpl.cell(row=tpl_row, column=col)
        dst = ws_out.cell(row=out_row, column=col)
        copy_style(src, dst)
    ws_out.row_dimensions[out_row].height = ws_tpl.row_dimensions[tpl_row].height or 45

def apply_day_border(ws, first_row, last_row, num_cols=12):
    """Kẻ thick border trên (đầu ngày) và dưới (cuối ngày) cho tất cả cột."""
    thick = Side(style='medium')   # medium = đậm rõ, không quá dày
    thin  = Side(style='thin')

    for row in range(first_row, last_row + 1):
        is_first = (row == first_row)
        is_last  = (row == last_row)
        for col in range(1, num_cols + 1):
            c = ws.cell(row=row, column=col)
            old = c.border
            top    = thick if is_first else (old.top    or thin)
            bottom = thick if is_last  else (old.bottom or thin)
            left   = old.left   or thin
            right  = old.right  or thin
            c.border = Border(top=top, bottom=bottom, left=left, right=right)


# ── Main builder ─────────────────────────────────────────────────
def build_from_template(report_data, year_month):
    y, m = year_month.split('-')

    # Load template
    tpl = load_workbook(TEMPLATE_PATH)
    ws_tpl = tpl['BAO_CAO']

    # Tạo workbook output mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{m}{y}'

    NUM_COLS = 12

    # ── Copy column widths từ template ──
    for col in range(1, NUM_COLS + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = ws_tpl.column_dimensions[letter].width or 15

    # ── ROW 1: Title — copy từ template, giữ merge A1:L1 ──
    ws.row_dimensions[1].height = ws_tpl.row_dimensions[1].height or 35
    ws.merge_cells('A1:L1')
    src = ws_tpl['A1']
    dst = ws['A1']
    dst.value = src.value
    copy_style(src, dst)

    # ── ROW 2: Headers — copy từng cell ──
    ws.row_dimensions[2].height = ws_tpl.row_dimensions[2].height or 60
    for col in range(1, NUM_COLS + 1):
        src = ws_tpl.cell(row=2, column=col)
        dst = ws.cell(row=2, column=col, value=src.value)
        copy_style(src, dst)

    # ── Sort dates, tính week ──
    dates = sorted(report_data.keys())
    week_totals = {1:0, 2:0, 3:0, 4:0}
    week_dates  = {1:[], 2:[], 3:[], 4:[]}
    for d in dates:
        wk = get_week(int(d.split('-')[2]))
        week_dates[wk].append(d)
        week_totals[wk] += sum(r['kpi'] for r in report_data[d])
    grand = sum(week_totals.values())

    week_labels = {
        1: 'TUẦN 1: (第1周):   ', 2: 'TUẦN 2: (第2周):   ',
        3: 'TUẦN 3: (第三周):   ', 4: 'TUẦN 4: (第四周):   '
    }

    # ── Write data rows (bắt đầu từ row 3) ──
    current_row    = 3
    week_first_row = {}
    week_last_row  = {}
    day_row_map    = {}   # date -> (first_row, last_row)

    for d in dates:
        rows = report_data[d]
        day  = int(d.split('-')[2])
        wk   = get_week(day)
        dt   = datetime(int(y), int(m), day)
        n    = len(rows)
        first_r = current_row

        if wk not in week_first_row:
            week_first_row[wk] = current_row

        for i, r in enumerate(rows):
            # Clone style từ dòng mẫu (row 3) của template
            copy_row_style(ws_tpl, 3, ws, current_row, NUM_COLS)

            is_empty_day = r.get('isEmpty', False)

            # A: Ngày — chỉ dòng đầu của ngày
            c = ws.cell(row=current_row, column=1)
            if i == 0:
                c.value = dt
                c.number_format = 'D/M'
            else:
                c.value = ''

            if is_empty_day:
                # Ngày trống: ghi x vào user, member, KPI — tổng = 0
                ws.cell(row=current_row, column=2).value = 'x'
                ws.cell(row=current_row, column=3).value = 'x'
                c = ws.cell(row=current_row, column=4)
                c.value = 'x'
            else:
                # B: User name
                ws.cell(row=current_row, column=2).value = r['userName']
                # C: Member
                ws.cell(row=current_row, column=3).value = r['memberName']
                # D: KPI ngày
                c = ws.cell(row=current_row, column=4)
                c.value = r['kpi']
                c.number_format = '#,##0'

            # E-G: trống (style đã copy)
            # H: x
            ws.cell(row=current_row, column=8).value = 'x'
            # I: KPI tổng ngày — chỉ dòng đầu, để trống còn lại
            c = ws.cell(row=current_row, column=9)
            day_kpi_total = sum(rr['kpi'] for rr in rows)
            c.value = day_kpi_total if i == 0 else ''
            if i == 0: c.number_format = '#,##0'
            # J: KPI tổng tuần — để trống trước, set sau khi biết first row tuần
            ws.cell(row=current_row, column=10).value = ''
            # K: Tuần label — để trống trước
            ws.cell(row=current_row, column=11).value = ''
            # L: Ghi chú — trống

            current_row += 1

        last_r = current_row - 1
        day_row_map[d]    = (first_r, last_r)
        week_last_row[wk] = last_r

        # Kẻ thick border đầu/cuối mỗi ngày
        apply_day_border(ws, first_r, last_r, NUM_COLS)

    # ── Merge cột A (Ngày) theo từng ngày ──
    for d, (r1, r2) in day_row_map.items():
        if r2 > r1:
            ws.merge_cells(f'A{r1}:A{r2}')
            ws.merge_cells(f'I{r1}:I{r2}')

    # ── Apply thick border phân cách từng ngày (SAU khi merge) ──
    # openpyxl merged cell chỉ render border từ top-left cell của vùng merge.
    # → Với cột A và I (merged per day): set cả top+bottom vào top-left cell (r1).
    # → Với các cột còn lại: set top tại r1, bottom tại r2 bình thường.
    thick = Side(style='medium')
    thin  = Side(style='thin')
    MERGED_COLS = {1, 9}   # cột A=1, I=9

    def safe(s):
        return s if (s and s.border_style) else thin

    for d, (r1, r2) in day_row_map.items():
        for col in range(1, NUM_COLS + 1):
            if col in MERGED_COLS and r2 > r1:
                # Merged cell: chỉ top-left (r1) có hiệu lực
                # → set top=thick và bottom=thick vào r1
                c = ws.cell(row=r1, column=col)
                b = c.border
                c.border = Border(top=thick, bottom=thick, left=safe(b.left), right=safe(b.right))
            else:
                # Non-merged: set top tại r1, bottom tại r2
                c_top = ws.cell(row=r1, column=col)
                b = c_top.border
                c_top.border = Border(top=thick, bottom=safe(b.bottom), left=safe(b.left), right=safe(b.right))

                c_bot = ws.cell(row=r2, column=col)
                b = c_bot.border
                c_bot.border = Border(top=safe(b.top), bottom=thick, left=safe(b.left), right=safe(b.right))

    # ── Set KPI tổng tuần + tuần label + merge J, K theo tuần ──
    for wk in [1, 2, 3, 4]:
        if wk not in week_first_row: continue
        r1, r2 = week_first_row[wk], week_last_row[wk]
        c = ws.cell(row=r1, column=10)
        c.value = week_totals[wk]
        c.number_format = '#,##0'
        ws.cell(row=r1, column=11).value = week_labels[wk]
        if r2 > r1:
            ws.merge_cells(f'J{r1}:J{r2}')
            ws.merge_cells(f'K{r1}:K{r2}')

    # ── Footer rows — copy style từ template row 4, 5 ──
    fr1, fr2 = current_row, current_row + 1
    ws.row_dimensions[fr1].height = ws_tpl.row_dimensions[4].height or 35
    ws.row_dimensions[fr2].height = ws_tpl.row_dimensions[5].height or 35

    # Row footer 1 (x)
    for col in range(1, NUM_COLS + 1):
        src = ws_tpl.cell(row=4, column=col)
        dst = ws.cell(row=fr1, column=col)
        copy_style(src, dst)
    ws.merge_cells(f'A{fr1}:H{fr1}')
    ws['A'+str(fr1)].value = 'x'
    ws.merge_cells(f'I{fr1}:L{fr1}')
    c = ws['I'+str(fr1)]
    c.value = grand
    c.number_format = '#,##0'
    copy_style(ws_tpl['I4'], c)

    # Row footer 2 (TỔNG KPI)
    for col in range(1, NUM_COLS + 1):
        src = ws_tpl.cell(row=5, column=col)
        dst = ws.cell(row=fr2, column=col)
        copy_style(src, dst)
    ws.merge_cells(f'A{fr2}:H{fr2}')
    ws['A'+str(fr2)].value = 'TỔNG KPI ĐỦ SỐ (KPI全部 - 达)'
    ws.merge_cells(f'I{fr2}:L{fr2}')
    c = ws['I'+str(fr2)]
    c.value = grand
    c.number_format = '#,##0'
    copy_style(ws_tpl['I5'], c)

    # ── Sheet 2: Data ──
    ws2 = wb.create_sheet('Data')
    ws2.append(['Date', 'UserName', 'MemberName', 'KPI', 'Week'])
    for d in dates:
        for r in report_data[d]:
            ws2.append([d, r['userName'], r['memberName'], r['kpi'],
                        f"Tuần {get_week(int(d.split('-')[2]))}"])
    for col, w in zip(['A','B','C','D','E'], [14, 25, 20, 12, 12]):
        ws2.column_dimensions[col].width = w

    return wb


if __name__ == '__main__':
    app.run(debug=True, port=5000)
